from typing import Dict, Any, List, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import json
from io import BytesIO
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from app.db.mongodb import db
from bson import ObjectId
import datetime

class TimetableExporter:
    """Export timetable data to various formats (Excel, PDF, JSON, CSV)"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
    
    async def export_timetable(self, timetable_id: str, format_type: str = "excel") -> BytesIO:
        """Export timetable in specified format"""
        try:
            # Get timetable data
            timetable = await self._get_timetable_data(timetable_id)
            
            if format_type.lower() == "excel":
                return await self._export_to_excel(timetable)
            elif format_type.lower() == "pdf":
                return await self._export_to_pdf(timetable)
            elif format_type.lower() == "json":
                return await self._export_to_json(timetable)
            elif format_type.lower() == "csv":
                return await self._export_to_csv(timetable)
            else:
                raise ValueError(f"Unsupported format: {format_type}")
                
        except Exception as e:
            raise Exception(f"Export failed: {str(e)}")
    
    async def _get_timetable_data(self, timetable_id: str) -> Dict[str, Any]:
        """Get comprehensive timetable data with related documents"""
        try:
            # Get timetable
            timetable = await db.db.timetables.find_one({"_id": ObjectId(timetable_id)})
            if not timetable:
                raise ValueError("Timetable not found")
            
            # Get program data
            program = await db.db.programs.find_one({"_id": timetable["program_id"]})
            
            # Get courses, faculty, and rooms for all entries
            entries_with_details = []
            for entry in timetable.get("entries", []):
                # Get course
                course = await db.db.courses.find_one({"_id": entry["course_id"]})
                # Get faculty
                faculty = await db.db.faculty.find_one({"_id": entry["faculty_id"]})
                # Get room
                room = await db.db.rooms.find_one({"_id": entry["room_id"]})
                
                entry_detail = {
                    "course_code": course.get("code", "N/A") if course else "N/A",
                    "course_name": course.get("name", "N/A") if course else "N/A",
                    "course_credits": course.get("credits", 0) if course else 0,
                    "faculty_name": faculty.get("name", "N/A") if faculty else "N/A",
                    "faculty_department": faculty.get("department", "N/A") if faculty else "N/A",
                    "room_number": room.get("number", "N/A") if room else "N/A",
                    "room_type": room.get("type", "N/A") if room else "N/A",
                    "room_capacity": room.get("capacity", 0) if room else 0,
                    "day": entry["time_slot"]["day"],
                    "start_time": entry["time_slot"]["start_time"],
                    "end_time": entry["time_slot"]["end_time"],
                    "duration": entry["time_slot"]["duration_minutes"],
                    "entry_type": entry.get("entry_type", "lecture")
                }
                entries_with_details.append(entry_detail)
            
            return {
                "timetable_id": str(timetable["_id"]),
                "program_name": program.get("name", "N/A") if program else "N/A",
                "program_code": program.get("code", "N/A") if program else "N/A",
                "semester": timetable.get("semester", "N/A"),
                "academic_year": timetable.get("academic_year", "N/A"),
                "created_at": timetable.get("created_at", datetime.datetime.utcnow()),
                "validation_status": timetable.get("validation_status", "pending"),
                "optimization_score": timetable.get("optimization_score", 0),
                "entries": entries_with_details
            }
            
        except Exception as e:
            raise Exception(f"Failed to get timetable data: {str(e)}")
    
    async def _export_to_excel(self, timetable_data: Dict[str, Any]) -> BytesIO:
        """Export timetable to Excel format"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Timetable"
            
            # Header styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title
            ws.merge_cells('A1:K1')
            ws['A1'] = f"Timetable - {timetable_data['program_name']}"
            ws['A1'].font = Font(bold=True, size=16)
            
            # Metadata
            ws['A3'] = f"Program: {timetable_data['program_code']}"
            ws['A4'] = f"Semester: {timetable_data['semester']}"
            ws['A5'] = f"Academic Year: {timetable_data['academic_year']}"
            ws['A6'] = f"Generated: {timetable_data['created_at'].strftime('%Y-%m-%d %H:%M')}"
            
            # Headers
            headers = [
                "Day", "Time", "Course Code", "Course Name", "Credits",
                "Faculty", "Department", "Room", "Room Type", "Capacity", "Type"
            ]
            
            row = 8
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            # Data rows
            for entry in timetable_data["entries"]:
                row += 1
                data = [
                    entry["day"],
                    f"{entry['start_time']} - {entry['end_time']}",
                    entry["course_code"],
                    entry["course_name"],
                    entry["course_credits"],
                    entry["faculty_name"],
                    entry["faculty_department"],
                    entry["room_number"],
                    entry["room_type"],
                    entry["room_capacity"],
                    entry["entry_type"].title()
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            # Save to BytesIO
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer
            
        except Exception as e:
            raise Exception(f"Excel export failed: {str(e)}")
    
    async def _export_to_pdf(self, timetable_data: Dict[str, Any]) -> BytesIO:
        """Export timetable to PDF format"""
        try:
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                                  leftMargin=0.5*inch, rightMargin=0.5*inch,
                                  topMargin=0.5*inch, bottomMargin=0.5*inch)
            
            elements = []
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=self.styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # Center
            )
            title = Paragraph(f"Timetable - {timetable_data['program_name']}", title_style)
            elements.append(title)
            
            # Metadata
            meta_style = ParagraphStyle(
                'MetaStyle',
                parent=self.styles['Normal'],
                fontSize=10,
                spaceAfter=6
            )
            
            meta_info = [
                f"Program: {timetable_data['program_code']}",
                f"Semester: {timetable_data['semester']}",
                f"Academic Year: {timetable_data['academic_year']}",
                f"Generated: {timetable_data['created_at'].strftime('%Y-%m-%d %H:%M')}",
                f"Status: {timetable_data['validation_status'].title()}"
            ]
            
            for info in meta_info:
                elements.append(Paragraph(info, meta_style))
            
            elements.append(Spacer(1, 20))
            
            # Table data
            table_data = []
            headers = ["Day", "Time", "Course", "Faculty", "Room", "Type"]
            table_data.append(headers)
            
            for entry in timetable_data["entries"]:
                row = [
                    entry["day"],
                    f"{entry['start_time']}-{entry['end_time']}",
                    f"{entry['course_code']}\n{entry['course_name'][:30]}",
                    entry["faculty_name"],
                    f"{entry['room_number']}\n({entry['room_type']})",
                    entry["entry_type"].title()
                ]
                table_data.append(row)
            
            # Create table
            table = Table(table_data, colWidths=[1*inch, 1.2*inch, 2*inch, 1.5*inch, 1*inch, 0.8*inch])
            
            # Table style
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            elements.append(table)
            
            # Build PDF
            doc.build(elements)
            buffer.seek(0)
            return buffer
            
        except Exception as e:
            raise Exception(f"PDF export failed: {str(e)}")
    
    async def _export_to_json(self, timetable_data: Dict[str, Any]) -> BytesIO:
        """Export timetable to JSON format"""
        try:
            # Convert datetime to string for JSON serialization
            export_data = timetable_data.copy()
            export_data["created_at"] = export_data["created_at"].isoformat()
            export_data["exported_at"] = datetime.datetime.utcnow().isoformat()
            
            json_str = json.dumps(export_data, indent=2, ensure_ascii=False)
            buffer = BytesIO(json_str.encode('utf-8'))
            buffer.seek(0)
            return buffer
            
        except Exception as e:
            raise Exception(f"JSON export failed: {str(e)}")
    
    async def _export_to_csv(self, timetable_data: Dict[str, Any]) -> BytesIO:
        """Export timetable to CSV format"""
        try:
            # Create DataFrame
            df = pd.DataFrame(timetable_data["entries"])
            
            # Reorder and rename columns
            column_mapping = {
                "day": "Day",
                "start_time": "Start Time",
                "end_time": "End Time",
                "course_code": "Course Code",
                "course_name": "Course Name",
                "course_credits": "Credits",
                "faculty_name": "Faculty",
                "faculty_department": "Department",
                "room_number": "Room",
                "room_type": "Room Type",
                "room_capacity": "Capacity",
                "entry_type": "Type"
            }
            
            df = df.rename(columns=column_mapping)
            df = df[list(column_mapping.values())]
            
            # Convert to CSV
            buffer = BytesIO()
            df.to_csv(buffer, index=False, encoding='utf-8')
            buffer.seek(0)
            return buffer
            
        except Exception as e:
            raise Exception(f"CSV export failed: {str(e)}")
    
    async def export_multiple_timetables(self, timetable_ids: List[str], format_type: str = "excel") -> BytesIO:
        """Export multiple timetables in a single file"""
        try:
            if format_type.lower() == "excel":
                return await self._export_multiple_to_excel(timetable_ids)
            elif format_type.lower() == "json":
                return await self._export_multiple_to_json(timetable_ids)
            else:
                raise ValueError(f"Multiple export not supported for format: {format_type}")
                
        except Exception as e:
            raise Exception(f"Multiple export failed: {str(e)}")
    
    async def _export_multiple_to_excel(self, timetable_ids: List[str]) -> BytesIO:
        """Export multiple timetables to Excel with separate sheets"""
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            for i, timetable_id in enumerate(timetable_ids):
                timetable_data = await self._get_timetable_data(timetable_id)
                
                # Create sheet
                ws = wb.create_sheet(title=f"Timetable_{i+1}")
                
                # Add data (simplified version of single export)
                ws['A1'] = f"{timetable_data['program_name']} - Sem {timetable_data['semester']}"
                
                headers = ["Day", "Time", "Course", "Faculty", "Room"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=3, column=col, value=header)
                
                for row, entry in enumerate(timetable_data["entries"], 4):
                    ws.cell(row=row, column=1, value=entry["day"])
                    ws.cell(row=row, column=2, value=f"{entry['start_time']}-{entry['end_time']}")
                    ws.cell(row=row, column=3, value=f"{entry['course_code']} - {entry['course_name']}")
                    ws.cell(row=row, column=4, value=entry["faculty_name"])
                    ws.cell(row=row, column=5, value=entry["room_number"])
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer
            
        except Exception as e:
            raise Exception(f"Multiple Excel export failed: {str(e)}")
    
    async def _export_multiple_to_json(self, timetable_ids: List[str]) -> BytesIO:
        """Export multiple timetables to JSON format"""
        try:
            all_timetables = []
            
            for timetable_id in timetable_ids:
                timetable_data = await self._get_timetable_data(timetable_id)
                timetable_data["created_at"] = timetable_data["created_at"].isoformat()
                all_timetables.append(timetable_data)
            
            export_data = {
                "exported_at": datetime.datetime.utcnow().isoformat(),
                "total_timetables": len(all_timetables),
                "timetables": all_timetables
            }
            
            json_str = json.dumps(export_data, indent=2, ensure_ascii=False)
            buffer = BytesIO(json_str.encode('utf-8'))
            buffer.seek(0)
            return buffer
            
        except Exception as e:
            raise Exception(f"Multiple JSON export failed: {str(e)}")